"""
For every supplier URL in suppliers.xlsx:
  1. Check {url}/robots.txt for a Sitemap: line
  2. If none, try the common default {url}/sitemap.xml
  3. Recursively walk sitemap indexes (a sitemap.xml that just lists
     other sitemap.xml files) down to real <url> entries
  4. Heuristically keep URLs that look like product/service pages
     (path contains "product", "service", "solutions", "offering", etc.)
  5. Use Claude to name each offering from its URL + page title
  6. Write Supplier Name / Type / Offering Name / URL rows into data.xlsx
     - Tags is left blank here; tag_offerings.py fills it in next.
  7. Any supplier with no usable sitemap is written with
     Type="NEEDS REVIEW" so it's easy to filter and check by hand.

Run:
    python crawl_offerings.py
"""

import re
import sys
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

sys.path.append(str(Path(__file__).resolve().parent.parent))
from config import SUPPLIERS_XLSX, OFFERINGS_XLSX, ANTHROPIC_MODEL  # noqa: E402

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; OfferingCrawlerBot/1.0)"}
REQUEST_TIMEOUT = 15
MAX_SITEMAPS_PER_SUPPLIER = 20
MAX_URLS_PER_SUPPLIER = 200

OFFERING_PATH_HINTS = re.compile(
    r"(product|products|service|services|solution|solutions|offering|catalog|shop|item)",
    re.IGNORECASE,
)


def find_sitemap_urls(base_url: str) -> list[str]:
    sitemaps = []
    try:
        robots = requests.get(urljoin(base_url, "/robots.txt"), headers=HEADERS, timeout=REQUEST_TIMEOUT)
        if robots.ok:
            for line in robots.text.splitlines():
                if line.lower().startswith("sitemap:"):
                    sitemaps.append(line.split(":", 1)[1].strip())
    except Exception:
        pass

    if not sitemaps:
        candidate = urljoin(base_url, "/sitemap.xml")
        try:
            resp = requests.head(candidate, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            if resp.ok:
                sitemaps.append(candidate)
        except Exception:
            pass

    return sitemaps


def expand_sitemaps(sitemap_urls: list[str]) -> list[str]:
    """Walk sitemap-index files down to plain <url> entries."""
    page_urls = []
    queue = list(sitemap_urls)
    seen = set()

    while queue and len(seen) < MAX_SITEMAPS_PER_SUPPLIER:
        sm_url = queue.pop(0)
        if sm_url in seen:
            continue
        seen.add(sm_url)

        try:
            resp = requests.get(sm_url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.content, "xml")
        except Exception as e:
            print(f"    ! could not read sitemap {sm_url}: {e}")
            continue

        nested = [loc.text.strip() for loc in soup.find_all("sitemap")]
        if nested:
            child_locs = [s.find("loc").text.strip() for s in soup.find_all("sitemap") if s.find("loc")]
            queue.extend(child_locs)
        else:
            page_urls.extend(loc.text.strip() for loc in soup.find_all("loc"))

        if len(page_urls) >= MAX_URLS_PER_SUPPLIER:
            break

    return page_urls[:MAX_URLS_PER_SUPPLIER]


def guess_offering_urls(all_urls: list[str]) -> list[str]:
    return [u for u in all_urls if OFFERING_PATH_HINTS.search(urlparse(u).path)]


def page_title(url: str) -> str | None:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")
        if soup.title and soup.title.string:
            return soup.title.string.strip()
        h1 = soup.find("h1")
        return h1.get_text(strip=True) if h1 else None
    except Exception:
        return None


def clean_offering_name(raw_title: str | None, url: str) -> str:
    if raw_title:
        # strip common "Title | Company Name" / "Title - Company" suffixes
        name = re.split(r"\s*[\|\-–]\s*", raw_title)[0].strip()
        if name:
            return name
    # fall back to the last non-empty path segment, prettified
    segment = [p for p in urlparse(url).path.split("/") if p][-1:]
    return segment[0].replace("-", " ").replace("_", " ").title() if segment else url


def load_existing_rows(path: Path) -> tuple[Workbook, list, dict]:
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "Supplier Name", "Type", "Offering Name", "Tags", "URL"])

    headers = [c.value for c in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    return wb, ws, col


def main():
    suppliers_wb = load_workbook(SUPPLIERS_XLSX)
    suppliers_ws = suppliers_wb.active
    s_headers = [c.value for c in suppliers_ws[1]]
    s_col = {name: idx + 1 for idx, name in enumerate(s_headers)}

    out_path = Path(OFFERINGS_XLSX)
    out_wb, out_ws, out_col = load_existing_rows(out_path)
    next_row = out_ws.max_row + 1

    for row in range(2, suppliers_ws.max_row + 1):
        name = suppliers_ws.cell(row=row, column=s_col["Supplier Name"]).value
        url = suppliers_ws.cell(row=row, column=s_col.get("URL", 0)).value if "URL" in s_col else None
        if not name or not url:
            continue

        print(f"\n{name} ({url})")
        sitemap_urls = find_sitemap_urls(url)

        if not sitemap_urls:
            print("  ! no sitemap found - flagging for review")
            out_ws.cell(row=next_row, column=out_col["Supplier Name"], value=name)
            out_ws.cell(row=next_row, column=out_col["Type"], value="NEEDS REVIEW")
            out_ws.cell(row=next_row, column=out_col["Offering Name"], value="(no sitemap - review manually)")
            out_ws.cell(row=next_row, column=out_col["URL"], value=url)
            next_row += 1
            continue

        all_urls = expand_sitemaps(sitemap_urls)
        offering_urls = guess_offering_urls(all_urls)
        print(f"  found {len(all_urls)} sitemap URLs, {len(offering_urls)} look like offerings")

        for o_url in offering_urls:
            title = page_title(o_url)
            offering_name = clean_offering_name(title, o_url)

            out_ws.cell(row=next_row, column=out_col["Supplier Name"], value=name)
            out_ws.cell(row=next_row, column=out_col["Type"], value="Product/Service")
            out_ws.cell(row=next_row, column=out_col["Offering Name"], value=offering_name)
            out_ws.cell(row=next_row, column=out_col["URL"], value=o_url)
            next_row += 1
            time.sleep(0.3)

    out_wb.save(out_path)
    print(f"\nDone. Wrote offerings to {out_path}")
    print("Next: run tag_offerings.py, then `python converter.py data`")


if __name__ == "__main__":
    main()
