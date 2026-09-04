"""
For every row in data.xlsx with an empty Tags column, ask Claude which
tags (from tags.xlsx) best fit the offering, based on its name/type and
- if reachable - the live page content. Writes a comma-separated list
into the Tags column.

Requires ANTHROPIC_API_KEY to be set in the environment.

Run:
    python tag_offerings.py
"""

import os
import sys
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parent.parent))
from config import OFFERINGS_XLSX, TAGS_XLSX, ANTHROPIC_MODEL  # noqa: E402

import anthropic  # noqa: E402

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; OfferingTaggerBot/1.0)"}


def load_tag_catalog() -> list[dict]:
    df = pd.read_excel(TAGS_XLSX)
    return df.to_dict(orient="records")


def page_snippet(url: str, max_chars: int = 1500) -> str:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")
        for tag in soup(["script", "style", "nav", "footer"]):
            tag.decompose()
        text = " ".join(soup.get_text(" ").split())
        return text[:max_chars]
    except Exception:
        return ""


def pick_tags(client: "anthropic.Anthropic", offering_name: str, offering_type: str,
              url: str, tag_catalog: list[dict]) -> str:
    catalog_text = "\n".join(
        f"- {t['Tag']}: {t.get('Keywords', '')} (Profession: {t.get('Profession', '')})"
        for t in tag_catalog
    )
    snippet = page_snippet(url)

    prompt = f"""You are tagging a catalog entry with the tags that best describe it.

Offering name: {offering_name}
Type: {offering_type}
URL: {url}
Page excerpt: {snippet or "(page not reachable, use the name/type/URL only)"}

Available tags (pick only from this list):
{catalog_text}

Return ONLY a comma-separated list of the tag names (from the list above)
that genuinely fit this offering. Use as few or as many as are accurate -
don't force a match. If nothing fits, return an empty string."""

    resp = client.messages.create(
        model=ANTHROPIC_MODEL,
        max_tokens=150,
        messages=[{"role": "user", "content": prompt}],
    )
    text = "".join(b.text for b in resp.content if b.type == "text").strip()
    return text


def main():
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("Set ANTHROPIC_API_KEY in your environment first.")
        sys.exit(1)

    client = anthropic.Anthropic()
    tag_catalog = load_tag_catalog()

    wb = load_workbook(OFFERINGS_XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}

    updated = 0
    for row in range(2, ws.max_row + 1):
        existing_tags = ws.cell(row=row, column=col["Tags"]).value
        if existing_tags:
            continue

        offering_name = ws.cell(row=row, column=col["Offering Name"]).value
        offering_type = ws.cell(row=row, column=col.get("Type", 0)).value or ""
        url = ws.cell(row=row, column=col.get("URL", 0)).value or ""
        if not offering_name:
            continue

        print(f"Row {row}: {offering_name}")
        tags = pick_tags(client, offering_name, offering_type, url, tag_catalog)
        ws.cell(row=row, column=col["Tags"], value=tags)
        print(f"  -> {tags or '(none)'}")
        updated += 1

    wb.save(OFFERINGS_XLSX)
    print(f"\nDone. Tagged {updated} offering(s). Saved to {OFFERINGS_XLSX}")
    print("Next: run `python converter.py data`")


if __name__ == "__main__":
    main()
