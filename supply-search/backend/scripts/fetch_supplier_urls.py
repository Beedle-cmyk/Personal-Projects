"""
For every row in suppliers.xlsx that has no URL yet:
  1. Search "{Supplier Name} {Supplier Address}"
  2. Take the first organic result
  3. Write it into the URL column
  4. Save the workbook (in place, with a .bak backup written first)

Search provider:
  Scraping Google's result pages directly breaks their Terms of Service
  and is fragile (selectors change, CAPTCHAs, rate limiting). Instead
  this script is written against a small `search_web()` interface with
  two backends:

    - "cse"  : Google Programmable Search Engine (Custom Search JSON API)
               - the sanctioned way to get Google results. Needs
               GOOGLE_CSE_API_KEY and GOOGLE_CSE_ID env vars.
               Free tier: 100 queries/day.
    - "ddg"  : DuckDuckGo HTML endpoint, used only as a no-signup
               fallback for light/manual use. Don't hammer it - add
               delay between requests (already built in below).

Set SEARCH_PROVIDER = "cse" once you have API keys; "ddg" works with
zero setup for smaller batches.

Run:
    python fetch_supplier_urls.py
"""

import os
import sys
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parent.parent))
from config import SUPPLIERS_XLSX  # noqa: E402

SEARCH_PROVIDER = os.environ.get("SEARCH_PROVIDER", "ddg")  # "cse" or "ddg"
DELAY_SECONDS = 1.5
HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; SupplierURLBot/1.0)"}


def search_web_cse(query: str) -> str | None:
    api_key = os.environ["GOOGLE_CSE_API_KEY"]
    cse_id = os.environ["GOOGLE_CSE_ID"]
    resp = requests.get(
        "https://www.googleapis.com/customsearch/v1",
        params={"key": api_key, "cx": cse_id, "q": query, "num": 1},
        timeout=15,
    )
    resp.raise_for_status()
    items = resp.json().get("items", [])
    return items[0]["link"] if items else None


def search_web_ddg(query: str) -> str | None:
    resp = requests.get(
        "https://html.duckduckgo.com/html/",
        params={"q": query},
        headers=HEADERS,
        timeout=15,
    )
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")
    link = soup.select_one("a.result__a")
    return link["href"] if link else None


def search_web(query: str) -> str | None:
    try:
        if SEARCH_PROVIDER == "cse":
            return search_web_cse(query)
        return search_web_ddg(query)
    except Exception as e:
        print(f"  ! search failed for '{query}': {e}")
        return None


def main():
    xlsx_path = Path(SUPPLIERS_XLSX)
    backup_path = xlsx_path.with_suffix(".xlsx.bak")
    if not backup_path.exists():
        backup_path.write_bytes(xlsx_path.read_bytes())
        print(f"Backed up original to {backup_path}")

    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}

    if "URL" not in col:
        ws.cell(row=1, column=len(headers) + 1, value="URL")
        col["URL"] = len(headers) + 1

    name_col = col["Supplier Name"]
    addr_col = col["Supplier Address"]
    url_col = col["URL"]

    updated = 0
    for row in range(2, ws.max_row + 1):
        existing_url = ws.cell(row=row, column=url_col).value
        if existing_url:
            continue

        name = ws.cell(row=row, column=name_col).value
        addr = ws.cell(row=row, column=addr_col).value
        if not name:
            continue

        query = f"{name} {addr or ''}".strip()
        print(f"Row {row}: searching '{query}'")
        url = search_web(query)
        if url:
            ws.cell(row=row, column=url_col, value=url)
            updated += 1
            print(f"  -> {url}")
        else:
            print("  -> no result found, leaving blank for manual review")

        time.sleep(DELAY_SECONDS)

    wb.save(xlsx_path)
    print(f"\nDone. Filled {updated} URL(s). Saved to {xlsx_path}")
    print("Next: run `python converter.py suppliers` to rebuild suppliers.db")


if __name__ == "__main__":
    main()
